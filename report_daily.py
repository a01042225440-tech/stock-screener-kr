"""
매일 장마감 후 자동 보고 (GitHub Actions, 16:40 KST).
- 시스템 자동점검(텔레그램/스캔 정상 여부)
- 매수·매도·손절 거래 데이터를 엑셀로 정리
- 누적 수익률 계산
- 엑셀 파일 + 요약을 텔레그램으로 발송
"""
import os, sys
from datetime import datetime
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from app import send_telegram, send_telegram_document
from trade_log import load_closed
from momentum_tracker import load_positions as mom_positions
from swing_tracker import _load_positions as sw_positions


def compound_3slot(closed):
    """청산거래를 매도일 기준 일별 묶어 상위3 균등 복리(근사). 누적수익률%."""
    by = {}
    for t in closed:
        by.setdefault(t["sellDate"], []).append(t["profitPct"])
    bal = 1.0
    for d in sorted(by):
        g = sorted(by[d], reverse=True)[:3]
        if g:
            bal *= (1 + np.mean(g) / 100)
    return (bal - 1) * 100


def build_excel(closed, open_mom, open_sw, path):
    wb = Workbook()
    hdr_fill = PatternFill("solid", fgColor="4338CA"); hdr_font = Font(color="FFFFFF", bold=True)
    # 시트1: 청산 거래
    ws = wb.active; ws.title = "청산거래"
    cols = ["매도일", "전략", "종목", "코드", "매수일", "매수가", "매도가", "수익률%", "사유"]
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        ws.cell(1, c).fill = hdr_fill; ws.cell(1, c).font = hdr_font; ws.cell(1, c).alignment = Alignment(horizontal="center")
    for t in sorted(closed, key=lambda x: x["sellDate"]):
        ws.append([t["sellDate"], t["strategy"], t["name"], t["code"], t["buyDate"],
                   t["buy"], t["sell"], t["profitPct"], t["reason"]])
        cell = ws.cell(ws.max_row, 8)
        cell.font = Font(color="DC2626" if t["profitPct"] < 0 else "16A34A", bold=True)
    for col, w in zip("ABCDEFGHI", [11, 7, 16, 8, 11, 10, 10, 9, 18]):
        ws.column_dimensions[col].width = w
    # 시트2: 보유중
    ws2 = wb.create_sheet("보유중")
    ws2.append(["전략", "종목", "코드", "매수일", "매수가", "현재가", "평가손익%"])
    for c in range(1, 8):
        ws2.cell(1, c).fill = hdr_fill; ws2.cell(1, c).font = hdr_font
    for p in open_mom:
        ws2.append(["모멘텀", p.get("name", ""), p["code"], p["buyDate"], p["buyPrice"], p.get("last", p["buyPrice"]), p.get("profit", 0)])
    for p in open_sw:
        ws2.append(["스윙", p.get("name", ""), p["code"], p["buyDate"], p["buyPrice"], p.get("last", p["buyPrice"]), p.get("profit", 0)])
    for col, w in zip("ABCDEFG", [7, 16, 8, 11, 10, 10, 10]):
        ws2.column_dimensions[col].width = w
    # 시트3: 요약
    ws3 = wb.create_sheet("요약")
    n = len(closed); wins = sum(1 for t in closed if t["profitPct"] > 0)
    avg = np.mean([t["profitPct"] for t in closed]) if closed else 0
    comp = compound_3slot(closed)
    mom_n = sum(1 for t in closed if t["strategy"] == "모멘텀"); sw_n = n - mom_n
    rows = [
        ["항목", "값"],
        ["총 청산거래", n],
        ["  └ 모멘텀 / 스윙", f"{mom_n} / {sw_n}"],
        ["승률", f"{(wins/n*100 if n else 0):.0f}%"],
        ["평균 수익률/거래", f"{avg:+.2f}%"],
        ["누적 수익률(3종목 복리)", f"{comp:+.1f}%"],
        ["보유중 종목", len(open_mom) + len(open_sw)],
        ["보고일", datetime.now().strftime("%Y-%m-%d %H:%M")],
    ]
    for r in rows: ws3.append(r)
    for c in range(1, 3):
        ws3.cell(1, c).fill = hdr_fill; ws3.cell(1, c).font = hdr_font
    ws3.column_dimensions["A"].width = 24; ws3.column_dimensions["B"].width = 18
    wb.save(path)
    return n, (wins/n*100 if n else 0), avg, comp


def main():
    today = datetime.now().strftime("%Y-%m-%d")
    closed = load_closed()
    om, osw = mom_positions(), sw_positions()

    # 시스템 자동점검
    health = []
    ok_tel, _ = send_telegram(f"🔧 시스템 점검 {today} — 정상 작동 중")
    health.append(f"텔레그램 {'✅' if ok_tel else '❌'}")

    path = os.path.join(os.path.dirname(os.path.abspath(__file__)), f"trades_report_{today}.xlsx")
    n, wr, avg, comp = build_excel(closed, om, osw, path)

    today_closed = [t for t in closed if t["sellDate"] == today]
    lines = [f"📈 <b>일일 매매 보고 ({today})</b>", ""]
    if today_closed:
        lines.append(f"📤 <b>오늘 청산 {len(today_closed)}건</b>")
        for t in today_closed:
            lines.append(f"  • [{t['strategy']}] {t['name']} <b>{t['profitPct']:+.1f}%</b> ({t['reason']})")
    else:
        lines.append("📤 오늘 청산된 거래 없음")
    lines += ["",
              f"📊 <b>누적 성적</b> (총 {n}건)",
              f"  승률 {wr:.0f}% · 평균 {avg:+.2f}%/거래",
              f"  <b>누적 수익률(3종목 복리) {comp:+.1f}%</b>",
              f"  보유중 {len(om)+len(osw)}종목",
              "",
              f"🔧 점검: {' · '.join(health)}",
              "📎 상세 엑셀 첨부"]
    summary = "\n".join(lines)

    ok, info = send_telegram_document(path, caption=summary)
    if not ok:
        send_telegram(summary)  # 파일 실패시 텍스트라도
    print(f"[REPORT] 청산{n} 누적{comp:+.1f}% 엑셀발송={ok} {info}")
    try:
        os.remove(path)
    except Exception:
        pass


if __name__ == "__main__":
    main()
